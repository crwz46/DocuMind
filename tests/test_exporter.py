from pathlib import Path

import pytest

from documind.exporter import ExcelExporter


class TestExcelExporter:
    def test_export_empty_data(self, tmp_path):
        output = str(tmp_path / "empty.xlsx")
        result = ExcelExporter.export([], output)
        assert Path(result).exists()
        assert Path(result).suffix == ".xlsx"

    def test_export_single_record(self, tmp_path):
        data = [{"name": "John", "age": 30, "city": "Jakarta"}]
        output = str(tmp_path / "single.xlsx")
        result = ExcelExporter.export(data, output)
        assert Path(result).exists()

    def test_export_multi_record(self, tmp_path):
        data = [
            {"name": "John", "age": 30},
            {"name": "Jane", "age": 25},
            {"name": "Bob", "age": 35},
        ]
        output = str(tmp_path / "multi.xlsx")
        result = ExcelExporter.export(data, output)
        assert Path(result).exists()

    def test_export_with_nested_data(self, tmp_path):
        data = [{"name": "John", "address": {"city": "Jakarta", "zip": "12345"}}]
        output = str(tmp_path / "nested.xlsx")
        result = ExcelExporter.export(data, output)
        assert Path(result).exists()

    def test_export_multi_sheet(self, tmp_path):
        data = {
            "Users": [{"name": "John", "age": 30}],
            "Products": [{"item": "Laptop", "price": 1500}],
        }
        output = str(tmp_path / "multi_sheet.xlsx")
        result = ExcelExporter.export_multi_sheet(data, output)
        assert Path(result).exists()

    def test_export_multi_sheet_empty(self, tmp_path):
        result = ExcelExporter.export_multi_sheet({"Empty": []}, str(tmp_path / "empty_multi.xlsx"))
        assert Path(result).exists()

    def test_file_is_valid_xlsx(self, tmp_path):
        data = [{"header1": "val1", "header2": "val2"}]
        output = str(tmp_path / "valid.xlsx")
        ExcelExporter.export(data, output)

        import openpyxl
        wb = openpyxl.load_workbook(output)
        ws = wb.active
        assert ws.title == "Extracted Data"
        assert ws.cell(1, 1).value == "header1"
        assert ws.cell(2, 1).value == "val1"
        wb.close()
