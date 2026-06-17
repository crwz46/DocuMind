from pathlib import Path
from typing import Any, Dict, List, Optional


class ExcelExporter:
    @staticmethod
    def export(data: List[Dict[str, Any]], output_path: str) -> str:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Extracted Data"

        if not data:
            ws.cell(row=1, column=1, value="No data extracted")
            wb.save(output_path)
            return output_path

        headers = list(data[0].keys())
        header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC"),
        )

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

        for row_idx, item in enumerate(data, 2):
            for col_idx, header in enumerate(headers, 1):
                val = item.get(header)
                if isinstance(val, (dict, list)):
                    val = str(val)
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top")

        for col_idx, header in enumerate(headers, 1):
            max_len = len(str(header))
            for row_idx in range(2, len(data) + 2):
                cell_val = ws.cell(row=row_idx, column=col_idx).value
                if cell_val:
                    max_len = max(max_len, len(str(cell_val)))
            ws.column_dimensions[chr(64 + col_idx)].width = min(max_len + 4, 60)

        ws.auto_filter.ref = ws.dimensions
        wb.save(output_path)
        return output_path

    @staticmethod
    def export_multi_sheet(data: Dict[str, List[Dict[str, Any]]], output_path: str) -> str:
        import openpyxl

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        for sheet_name, sheet_data in data.items():
            ws = wb.create_sheet(title=sheet_name[:31])
            if not sheet_data:
                ws.cell(row=1, column=1, value="No data")
                continue

            headers = list(sheet_data[0].keys())
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True)

            for row_idx, item in enumerate(sheet_data, 2):
                for col_idx, header in enumerate(headers, 1):
                    val = item.get(header)
                    if isinstance(val, (dict, list)):
                        val = str(val)
                    ws.cell(row=row_idx, column=col_idx, value=val)

        wb.save(output_path)
        return output_path
